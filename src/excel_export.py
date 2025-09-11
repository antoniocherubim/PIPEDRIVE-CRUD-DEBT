"""
Sistema de exportação para planilha de importação do Pipedrive
Gera arquivos Excel com colunas formatadas para importação direta
"""
import pandas as pd
import logging
from datetime import datetime
from typing import Dict, List, Any, Optional
import os

logger = logging.getLogger(__name__)

class PipedriveExcelExporter:
    """Exporta dados para planilha de importação do Pipedrive"""
    
    def __init__(self):
        self.output_folder = "output/excel_export"
        self._ensure_output_folder()
    
    def _ensure_output_folder(self):
        """Cria pasta de saída se não existir"""
        if not os.path.exists(self.output_folder):
            os.makedirs(self.output_folder, exist_ok=True)
    
    def export_inadimplentes_to_excel(self, inadimplentes_data: List[Dict], 
                                    filename: str = None) -> str:
        """
        Exporta inadimplentes para planilha de importação do Pipedrive
        
        Args:
            inadimplentes_data: Lista de dados dos inadimplentes
            filename: Nome do arquivo (opcional)
            
        Returns:
            Caminho do arquivo gerado
        """
        if not inadimplentes_data:
            raise ValueError("Nenhum dado fornecido para exportação")
        
        logger.info(f"Iniciando exportação de {len(inadimplentes_data)} inadimplentes para Excel")
        
        # Gerar nome do arquivo se não fornecido
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"pipedrive_import_{timestamp}.xlsx"
        
        # Garantir extensão .xlsx
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        filepath = os.path.join(self.output_folder, filename)
        
        try:
            # Preparar dados para exportação
            export_data = self._prepare_export_data(inadimplentes_data)
            
            # Criar DataFrame
            df = pd.DataFrame(export_data)
            
            # Salvar em Excel com formatação
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Aba principal com todos os dados
                df.to_excel(writer, sheet_name='Importação Pipedrive', index=False)
                
                # Aba de pessoas apenas
                pessoas_df = self._create_pessoas_dataframe(inadimplentes_data)
                pessoas_df.to_excel(writer, sheet_name='Pessoas', index=False)
                
                # Aba de negócios apenas
                negocios_df = self._create_negocios_dataframe(inadimplentes_data)
                negocios_df.to_excel(writer, sheet_name='Negócios', index=False)
                
                # Aplicar formatação
                self._apply_excel_formatting(writer)
            
            logger.info(f"Arquivo Excel gerado com sucesso: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Erro ao gerar arquivo Excel: {e}")
            raise
    
    def _prepare_export_data(self, inadimplentes_data: List[Dict]) -> List[Dict]:
        """Prepara dados para exportação combinando pessoa e negócio"""
        export_data = []
        
        for inadimplente in inadimplentes_data:
            # Dados base
            cpf_cnpj = inadimplente.get('cpf_cnpj', '')
            nome = inadimplente.get('nome', '')
            tipo_pessoa = inadimplente.get('tipo_pessoa', '')
            
            if not cpf_cnpj or tipo_pessoa == 'INDEFINIDO':
                continue
            
            # Criar registro combinado
            record = {
                # Identificação
                'ID_REFERENCIA': cpf_cnpj,
                'TIPO_PESSOA': tipo_pessoa,
                
                # Campos de Pessoa
                'Pessoa - Nome': nome,
                'Pessoa - CPF/CNPJ': cpf_cnpj,
                'Pessoa - Telefone': inadimplente.get('telefone_principal', ''),
                'Pessoa - Email': inadimplente.get('email_principal', ''),
                'Pessoa - Endereço': inadimplente.get('endereco_completo', ''),
                'Pessoa - Data Nascimento': inadimplente.get('data_nascimento', ''),
                'Pessoa - Estado Civil': inadimplente.get('estado_civil', ''),
                'Pessoa - Condição CPF': inadimplente.get('condicao_cpf', ''),
                'Pessoa - Nome da Mãe': inadimplente.get('nome_mae', ''),
                'Pessoa - Nome do Cônjuge': inadimplente.get('nome_conjuge', ''),
                
                # Campos de Negócio
                'Negócio - Título': f"Inadimplência - {nome}",
                'Negócio - Valor Total': inadimplente.get('valor_total_divida', ''),
                'Negócio - Valor Vencido': inadimplente.get('valor_total_vencido', ''),
                'Negócio - Dias de Atraso': inadimplente.get('dias_atraso', ''),
                'Negócio - Total de Parcelas': inadimplente.get('total_parcelas', ''),
                'Negócio - Vencimento Mais Antigo': inadimplente.get('vencimento_mais_antigo', ''),
                'Negócio - Data Prejuízo Mais Antigo': inadimplente.get('data_prejuizo_mais_antigo', ''),
                'Negócio - Data Terceirização': inadimplente.get('data_terceirizacao', ''),
                'Negócio - Data Prevista Honra': inadimplente.get('data_prevista_honra', ''),
                'Negócio - Valor Mínimo': inadimplente.get('valor_minimo', ''),
                'Negócio - Contrato Garantinorte': inadimplente.get('contrato_garantinorte', ''),
                'Negócio - Tag Atraso': inadimplente.get('tag_atraso', ''),
                'Negócio - Cooperado': inadimplente.get('cooperado', ''),
                'Negócio - Cooperativa': inadimplente.get('cooperativa', ''),
                'Negócio - Todos Contratos': inadimplente.get('todos_contratos', ''),
                'Negócio - Todas Operações': inadimplente.get('todas_operacoes', ''),
                'Negócio - Número Contrato': inadimplente.get('numero_contrato', ''),
                'Negócio - Tipo Ação Carteira': inadimplente.get('tipo_acao_carteira', ''),
                'Negócio - Avalistas': inadimplente.get('avalistas', ''),
                
                # Campos adicionais
                'DATA_EXPORTACAO': datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                'STATUS_PROCESSAMENTO': 'Pendente'
            }
            
            # Limpar campos vazios ou None
            record = {k: v for k, v in record.items() if v is not None and str(v).strip() != ''}
            
            export_data.append(record)
        
        return export_data
    
    def _create_pessoas_dataframe(self, inadimplentes_data: List[Dict]) -> pd.DataFrame:
        """Cria DataFrame específico para pessoas"""
        pessoas_data = []
        
        for inadimplente in inadimplentes_data:
            cpf_cnpj = inadimplente.get('cpf_cnpj', '')
            tipo_pessoa = inadimplente.get('tipo_pessoa', '')
            
            if not cpf_cnpj or tipo_pessoa == 'INDEFINIDO':
                continue
            
            pessoa = {
                'Pessoa - Nome': inadimplente.get('nome', ''),
                'Pessoa - CPF/CNPJ': cpf_cnpj,
                'Pessoa - Telefone': inadimplente.get('telefone_principal', ''),
                'Pessoa - Email': inadimplente.get('email_principal', ''),
                'Pessoa - Endereço': inadimplente.get('endereco_completo', ''),
                'Pessoa - Data Nascimento': inadimplente.get('data_nascimento', ''),
                'Pessoa - Estado Civil': inadimplente.get('estado_civil', ''),
                'Pessoa - Condição CPF': inadimplente.get('condicao_cpf', ''),
                'Pessoa - Nome da Mãe': inadimplente.get('nome_mae', ''),
                'Pessoa - Nome do Cônjuge': inadimplente.get('nome_conjuge', ''),
                'Pessoa - RG': inadimplente.get('rg', ''),
                'Pessoa - Data Emissão RG': inadimplente.get('data_emissao_rg', ''),
                'Pessoa - Órgão Emissor RG': inadimplente.get('orgao_emissor_rg', ''),
                'Pessoa - UF RG': inadimplente.get('uf_rg', ''),
                'TIPO_PESSOA': tipo_pessoa,
                'ID_REFERENCIA': cpf_cnpj
            }
            
            # Limpar campos vazios
            pessoa = {k: v for k, v in pessoa.items() if v is not None and str(v).strip() != ''}
            pessoas_data.append(pessoa)
        
        return pd.DataFrame(pessoas_data)
    
    def _create_negocios_dataframe(self, inadimplentes_data: List[Dict]) -> pd.DataFrame:
        """Cria DataFrame específico para negócios"""
        negocios_data = []
        
        for inadimplente in inadimplentes_data:
            cpf_cnpj = inadimplente.get('cpf_cnpj', '')
            nome = inadimplente.get('nome', '')
            tipo_pessoa = inadimplente.get('tipo_pessoa', '')
            
            if not cpf_cnpj or tipo_pessoa == 'INDEFINIDO':
                continue
            
            negocio = {
                'Negócio - Título': f"Inadimplência - {nome}",
                'Negócio - Pessoa (CPF/CNPJ)': cpf_cnpj,
                'Negócio - Valor Total': inadimplente.get('valor_total_divida', ''),
                'Negócio - Valor Vencido': inadimplente.get('valor_total_vencido', ''),
                'Negócio - Valor Total com Juros': inadimplente.get('valor_total_com_juros', ''),
                'Negócio - Dias de Atraso': inadimplente.get('dias_atraso', ''),
                'Negócio - Total de Parcelas': inadimplente.get('total_parcelas', ''),
                'Negócio - Vencimento Mais Antigo': inadimplente.get('vencimento_mais_antigo', ''),
                'Negócio - Data Prejuízo Mais Antigo': inadimplente.get('data_prejuizo_mais_antigo', ''),
                'Negócio - Data Terceirização': inadimplente.get('data_terceirizacao', ''),
                'Negócio - Data Prevista Honra': inadimplente.get('data_prevista_honra', ''),
                'Negócio - Valor Mínimo': inadimplente.get('valor_minimo', ''),
                'Negócio - Contrato Garantinorte': inadimplente.get('contrato_garantinorte', ''),
                'Negócio - Tag Atraso': inadimplente.get('tag_atraso', ''),
                'Negócio - Cooperado': inadimplente.get('cooperado', ''),
                'Negócio - Cooperativa': inadimplente.get('cooperativa', ''),
                'Negócio - Todos Contratos': inadimplente.get('todos_contratos', ''),
                'Negócio - Todas Operações': inadimplente.get('todas_operacoes', ''),
                'Negócio - Número Contrato': inadimplente.get('numero_contrato', ''),
                'Negócio - Tipo Ação Carteira': inadimplente.get('tipo_acao_carteira', ''),
                'Negócio - Avalistas': inadimplente.get('avalistas', ''),
                'Negócio - Especulação Parcelamento': inadimplente.get('especulacao_parcelamento', ''),
                'PIPELINE': 'BASE NOVA - SDR',  # Pipeline padrão
                'STAGE': 'Novo Lead',  # Stage padrão
                'ID_REFERENCIA': cpf_cnpj
            }
            
            # Limpar campos vazios
            negocio = {k: v for k, v in negocio.items() if v is not None and str(v).strip() != ''}
            negocios_data.append(negocio)
        
        return pd.DataFrame(negocios_data)
    
    def _apply_excel_formatting(self, writer):
        """Aplica formatação ao arquivo Excel"""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Cores
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(color='FFFFFF', bold=True)
            
            # Bordas
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Aplicar formatação para cada aba
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                
                # Formatar cabeçalho
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                
                # Auto-ajustar largura das colunas
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)  # Limitar a 50 caracteres
                    worksheet.column_dimensions[column_letter].width = adjusted_width
                
                # Aplicar bordas a todas as células com dados
                for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
                    for cell in row:
                        if cell.value is not None:
                            cell.border = thin_border
                
                # Congelar primeira linha
                worksheet.freeze_panes = 'A2'
                
        except ImportError:
            logger.warning("openpyxl não disponível para formatação avançada")
        except Exception as e:
            logger.warning(f"Erro ao aplicar formatação Excel: {e}")
    
    def export_custom_fields_template(self, filename: str = None) -> str:
        """
        Exporta template de campos personalizados do Pipedrive
        
        Returns:
            Caminho do arquivo template gerado
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"pipedrive_custom_fields_template_{timestamp}.xlsx"
        
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        filepath = os.path.join(self.output_folder, filename)
        
        try:
            # Dados dos campos personalizados
            custom_fields_data = [
                # Campos de Pessoa
                {
                    'TIPO': 'Pessoa',
                    'NOME_CAMPO': 'CPF',
                    'ID_CAMPO': 'SUBSTITUIR_PELO_ID_REAL',
                    'TIPO_CAMPO': 'Text',
                    'DESCRICAO': 'CPF da pessoa'
                },
                {
                    'TIPO': 'Pessoa',
                    'NOME_CAMPO': 'DATA_NASCIMENTO',
                    'ID_CAMPO': 'SUBSTITUIR_PELO_ID_REAL',
                    'TIPO_CAMPO': 'Date',
                    'DESCRICAO': 'Data de nascimento'
                },
                {
                    'TIPO': 'Pessoa',
                    'NOME_CAMPO': 'ESTADO_CIVIL',
                    'ID_CAMPO': 'SUBSTITUIR_PELO_ID_REAL',
                    'TIPO_CAMPO': 'Single Option',
                    'DESCRICAO': 'Estado civil'
                },
                {
                    'TIPO': 'Pessoa',
                    'NOME_CAMPO': 'CONDICAO_CPF',
                    'ID_CAMPO': 'SUBSTITUIR_PELO_ID_REAL',
                    'TIPO_CAMPO': 'Text',
                    'DESCRICAO': 'Condição do CPF'
                },
                {
                    'TIPO': 'Pessoa',
                    'NOME_CAMPO': 'ENDERECO',
                    'ID_CAMPO': 'SUBSTITUIR_PELO_ID_REAL',
                    'TIPO_CAMPO': 'Address',
                    'DESCRICAO': 'Endereço completo'
                },
                
                # Campos de Negócio
                {
                    'TIPO': 'Negócio',
                    'NOME_CAMPO': 'ID_CPF_CNPJ',
                    'ID_CAMPO': 'SUBSTITUIR_PELO_ID_REAL',
                    'TIPO_CAMPO': 'Numeric',
                    'DESCRICAO': 'ID do CPF/CNPJ'
                },
                {
                    'TIPO': 'Negócio',
                    'NOME_CAMPO': 'VALOR_TOTAL_DA_DIVIDA',
                    'ID_CAMPO': 'SUBSTITUIR_PELO_ID_REAL',
                    'TIPO_CAMPO': 'Monetary',
                    'DESCRICAO': 'Valor total da dívida'
                },
                {
                    'TIPO': 'Negócio',
                    'NOME_CAMPO': 'VALOR_TOTAL_VENCIDO',
                    'ID_CAMPO': 'SUBSTITUIR_PELO_ID_REAL',
                    'TIPO_CAMPO': 'Monetary',
                    'DESCRICAO': 'Valor total vencido'
                },
                {
                    'TIPO': 'Negócio',
                    'NOME_CAMPO': 'DIAS_DE_ATRASO',
                    'ID_CAMPO': 'SUBSTITUIR_PELO_ID_REAL',
                    'TIPO_CAMPO': 'Numeric',
                    'DESCRICAO': 'Dias de atraso'
                },
                {
                    'TIPO': 'Negócio',
                    'NOME_CAMPO': 'TAG_ATRASO',
                    'ID_CAMPO': 'SUBSTITUIR_PELO_ID_REAL',
                    'TIPO_CAMPO': 'Multiple Options',
                    'DESCRICAO': 'Tag de atraso'
                },
                {
                    'TIPO': 'Negócio',
                    'NOME_CAMPO': 'CONTRATO_GARANTINORTE',
                    'ID_CAMPO': 'SUBSTITUIR_PELO_ID_REAL',
                    'TIPO_CAMPO': 'Text',
                    'DESCRICAO': 'Contrato Garantinorte'
                }
            ]
            
            # Criar DataFrame
            df = pd.DataFrame(custom_fields_data)
            
            # Salvar em Excel
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='Campos Personalizados', index=False)
                
                # Aplicar formatação
                self._apply_excel_formatting(writer)
            
            logger.info(f"Template de campos personalizados gerado: {filepath}")
            return filepath
            
        except Exception as e:
            logger.error(f"Erro ao gerar template: {e}")
            raise
    
    def get_export_summary(self, filepath: str) -> Dict[str, Any]:
        """
        Retorna resumo da exportação
        
        Args:
            filepath: Caminho do arquivo exportado
            
        Returns:
            Dicionário com informações do arquivo
        """
        try:
            if not os.path.exists(filepath):
                return {'error': 'Arquivo não encontrado'}
            
            # Ler arquivo Excel
            df = pd.read_excel(filepath, sheet_name='Importação Pipedrive')
            
            # Contar registros
            total_records = len(df)
            
            # Contar por tipo de pessoa
            tipos_pessoa = df['TIPO_PESSOA'].value_counts().to_dict()
            
            # Informações do arquivo
            file_stats = os.stat(filepath)
            
            summary = {
                'arquivo': os.path.basename(filepath),
                'caminho_completo': filepath,
                'total_registros': total_records,
                'tipos_pessoa': tipos_pessoa,
                'tamanho_bytes': file_stats.st_size,
                'tamanho_mb': round(file_stats.st_size / (1024 * 1024), 2),
                'data_criacao': datetime.fromtimestamp(file_stats.st_ctime).strftime("%d/%m/%Y %H:%M:%S"),
                'colunas': list(df.columns)
            }
            
            return summary
            
        except Exception as e:
            logger.error(f"Erro ao gerar resumo: {e}")
            return {'error': str(e)}
